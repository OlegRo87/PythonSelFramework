import openpyxl

Dict = {}
book = openpyxl.load_workbook("/home/oleg/Downloads/udemy/automation/excel_automation.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Oleg"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet['A3'].value)

print('**********************')
# whole table
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value)
print('**********************')

# part of  table
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)
print('**********************')
# Dict
# part of  table
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column + 1):
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)
