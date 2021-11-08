import openpyxl


class HomePageData:
    test_home_page_data = [{"first_name": "oleg", "email": "oleg@gmail.com", "gender": "Male"},
                           {"first_name": "ilona", "email": "ilona@gmail.com", "gender": "Female"}]

    @staticmethod
    def get_test_data(testcase_name):
        dict = {}
        book = openpyxl.load_workbook("/home/oleg/Downloads/udemy/automation/excel_automation.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [dict]
